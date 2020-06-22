"""
find_missing_vars.py - Scan the Excel file generated directly from HighResMIP
in the latest data request and identify any variables that aren't in the 
current PRIMAVERA data request.

Crown Copyright (2017)
"""
from openpyxl import load_workbook


HIGHRESMIP = 'xls/cmvme_HighResMIP_highresSST-present_1_3.xlsx'
PRIMAVERA = 'PRIMAVERA_Data_Request_v01_00_07.xlsx'


def main():
    CMOR_NAME_COLUMN = 11
    """ Run the software """
    # EdayZ, fx removed
    sheet_names = [
        'Amon', 'LImon', 'Lmon', 'Omon', 'SImon', 'AERmon', 'CFmon', 'Emon',
        'EmonZ', 'Oday', 'CFday', 'day', 'Eday', 'SIday', '6hrPlev',
        '6hrPlevPt', '3hr', 'E3hr', 'E3hrPt', 'E1hr', 'Esubhr'
    ]
    highresmip = load_workbook(HIGHRESMIP)
    primavera = load_workbook(PRIMAVERA)

    for sheet in sheet_names:
        hrm_sheet = highresmip[sheet]
        prim_sheet = primavera[sheet]

        hrm_vars = [row[CMOR_NAME_COLUMN].value
                    for row in hrm_sheet.iter_rows(min_row=2)
                    if row[CMOR_NAME_COLUMN].value]

        prim_vars = [row[CMOR_NAME_COLUMN].value
                     for row in prim_sheet.iter_rows(min_row=2)
                     if row[CMOR_NAME_COLUMN].value]

        for cmor_var in hrm_vars:
            if cmor_var not in prim_vars:
                print '{} {} in HighResMIP but not PRIMAVERA'.format(
                    sheet, cmor_var)

        for cmor_var in prim_vars:
            if cmor_var not in hrm_vars:
                print '{} {} in PRIMAVERA but not HighResMIP'.format(
                    sheet, cmor_var)


if __name__ == '__main__':
    main()
