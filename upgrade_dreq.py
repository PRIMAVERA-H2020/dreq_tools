#!/usr/bin/env python2.7
"""
upgrade_dreq.py - upgrade the Excel copy of the data request to the latest
version of the data request.

Crown Copyright (2017)
"""

import simple_dreq as sdq
import openpyxl as xl
from openpyxl.styles import PatternFill

sdq.initialise()

INPUT_XLS = 'PRIMAVERA_Data_Request_v01_00_07.xlsx'
OUTPUT_XLS = 'PRIMAVERA_Data_Request_v01_00_13.xlsx'

print "Load the excel file"
workbook=xl.load_workbook(INPUT_XLS)

def unhighlight_cell(cell):
    cell_font = cell.font.copy()
    cell_font.bold = False
    cell_font.italic = False
    cell.font = cell_font
    cell.fill = PatternFill()

def highlight_cell(cell):
    cell_font = cell.font.copy()
    cell_font.bold = True
    cell.font = cell_font
    cell.fill = PatternFill(start_color= 'ffff00', end_color= 'ffff00', fill_type="solid")

#this may be useful or may not
#DREQ_COLUMNS = {"A": "Priority",
#                "B": "Long name",*
#                "C": "units",*
#                "D": "description",*
#                "E": "comment",*
#                "F": "Variable Name",*
#                "G": "CF Standard Name",*
#                "H": "cell_methods",*
#                "I": "positive",*
#                "J": "type",*
#                "K": "dimensions",*
#                "L": "CMOR Name",*
#                "M": "modeling_realm",*
#                "N": "frequency",*
#                "O": "cell_measures",*
#                "P": "prov",
#                "Q": "provNote",
#                "R": "rowIndex",
#                "S": "UID",*
#                "T": "vid",*
#                "U": "stid",*
#                "V": "Structure Title",
#                "W": "rubbish",
#                "X": "rubbish",
#                "Y": "rubbish",
#                "Z": "rubbish",
#                "AA": "MIPs (requesting)",* These are shifted over several columns
#                "AB": "MIPs (by experiment)",*
#                }

print "Un-bold current bolded stuff"

# Un-bold any cells with text in bold apart from the Notes page
for sheet in workbook.worksheets[1:]:
    for myrow in range(2, sheet.max_row+1):
        for mycol in range(1,29):
            unhighlight_cell(sheet.cell(row=myrow, column=mycol))

# Check for differences
for sheet in workbook.worksheets[1:]: #skip "Notes"
    print "------------------------------"
    print sheet
    for row in range(2, sheet.max_row+1):
        # CMOR name is the "primary key"
        cmor_name = sheet["L%i" % row].value
        uid = sdq.lookup_uid(cmor_name, sheet.title)

        first = 1

        if uid is None:
           print "  !! NO DATA REQUEST ENTRY FOR ", sheet.title, cmor_name,"!!"
           continue
        
       # Priority
        if sheet["A%i"  % row].value is None:
            current_pri = ""
        else:
            current_pri = sheet["A%i"  % row].value
        new_pri = sdq.get_priority(uid)


        if str(current_pri) != str(new_pri):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_pri=",current_pri,"to new_pri=",new_pri
            #raw_input("Press Enter to continue...")

            sheet["A%i"  % row].value = new_pri
            highlight_cell(sheet["A%i" % row])

       # Long Name
        if sheet["B%i"  % row].value is None:
            current_long_name = ""
        else:
            current_long_name = sheet["B%i"  % row].value
        new_long_name = sdq.get_long_name(uid)

        # work out if change is significant enough to highlight
        # significant_change = False

        if str(current_long_name) != str(new_long_name):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_long_name=",current_long_name,"to new_long_name=",new_long_name
            #raw_input("Press Enter to continue...")
            sheet["B%i"  % row].value = new_long_name
            highlight_cell(sheet["B%i" % row])

       # Units
        if sheet["C%i"  % row].value is None:
            current_unit = ""
        else:
            current_unit = sheet["C%i"  % row].value
        new_unit = sdq.get_units(uid)

        if str(current_unit) != str(new_unit):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_unit=",current_unit,"to new_unit=",new_unit
            #raw_input("Press Enter to continue...")
            sheet["C%i"  % row].value = new_unit
            highlight_cell(sheet["C%i" % row])

       # Description
        if sheet["D%i"  % row].value is None:
            current_desc = ""
        else:
            current_desc = sheet["D%i"  % row].value
        new_desc = sdq.get_description(uid)

        if str(current_desc) != str(new_desc):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_desc=",current_desc,"to new_desc=",new_desc
            #raw_input("Press Enter to continue...")
            sheet["D%i"  % row].value = new_desc
            highlight_cell(sheet["D%i" % row])

       # Comment
        if sheet["E%i"  % row].value is None:
            current_comm = ""
        else:
            current_comm = sheet["E%i"  % row].value
        new_comm = sdq.get_comment(uid)

        if str(current_comm) != str(new_comm):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_comm=",current_comm,"to new_comm=",new_comm
            #raw_input("Press Enter to continue...")
            sheet["E%i"  % row].value = new_comm
            highlight_cell(sheet["E%i" % row])

       # Variable Name
        if sheet["F%i"  % row].value is None:
            current_varname = ""
        else:
            current_varname = sheet["F%i"  % row].value
        new_varname = sdq.get_var_name(uid)

        if str(current_varname) != str(new_varname):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_varname=",current_varname,"to new_varname=",new_varname
            #raw_input("Press Enter to continue...")
            sheet["F%i"  % row].value = new_varname
            highlight_cell(sheet["F%i" % row])

       # CF standard name
        if sheet["G%i"  % row].value is None:
            current_cf = ""
        else:
            current_cf = sheet["G%i"  % row].value
        new_cf = sdq.get_standard_name(uid)

        if str(current_cf) != str(new_cf):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_cf=",current_cf,"to new_cf=",new_cf
            #raw_input("Press Enter to continue...")
            sheet["G%i"  % row].value = new_cf
            highlight_cell(sheet["G%i" % row])

       # Cell Methods
        if sheet["H%i"  % row].value is None:
            current_cellmethod = ""
        else:
            current_cellmethod = sheet["H%i"  % row].value
        new_cellmethod = sdq.get_cell_methods(uid)

        if str(current_cellmethod) != str(new_cellmethod):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_cellmethod=",current_cellmethod,"to new_cellmethod=",new_cellmethod
            #raw_input("Press Enter to continue...")
            sheet["H%i"  % row].value = new_cellmethod
            highlight_cell(sheet["H%i" % row])

       # Positive
        if sheet["I%i"  % row].value is None:
            current_pos = ""
        else:
            current_pos = sheet["I%i"  % row].value
        new_pos = sdq.get_positive(uid)

        if str(current_pos) != str(new_pos):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_pos=",current_pos,"to new_pos=",new_pos
            #raw_input("Press Enter to continue...")
            sheet["I%i"  % row].value = new_pos
            highlight_cell(sheet["I%i" % row])

       # Type
        if sheet["J%i"  % row].value is None:
            current_typ = ""
        else:
            current_typ = sheet["J%i"  % row].value
        new_typ = sdq.get_type(uid)

        if str(current_typ) != str(new_typ):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_typ=",current_typ,"to new_typ=",new_typ
            #raw_input("Press Enter to continue...")
            sheet["J%i"  % row].value = new_typ
            highlight_cell(sheet["J%i" % row])

       # Dimensions
        if sheet["K%i"  % row].value is None:
            current_dims = ""
        else:
            current_dims = sheet["K%i"  % row].value
        new_dims = sdq.get_dimensions(uid)

        if str(current_dims) != str(new_dims):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_dims=",current_dims,"to new_dims=",new_dims
            #raw_input("Press Enter to continue...")
            sheet["K%i"  % row].value = new_dims
            highlight_cell(sheet["K%i" % row])

       # modelling realm
        if sheet["M%i"  % row].value is None:
            current_mrealm = ""
        else:
            current_mrealm = sheet["M%i"  % row].value
        new_mrealm = sdq.get_modeling_realm(uid)

        if str(current_mrealm) != str(new_mrealm):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_mrealm=",current_mrealm,"to new_mrealm=",new_mrealm
            #raw_input("Press Enter to continue...")
            sheet["M%i"  % row].value = new_mrealm
            highlight_cell(sheet["M%i" % row])

       # Frequency
        if sheet["N%i"  % row].value is None:
            current_freq = ""
        else:
            current_freq = sheet["N%i"  % row].value
        new_freq = sdq.get_frequency(uid)

        if str(current_freq) != str(new_freq):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_freq=",current_freq,"to new_freq=",new_freq
            #raw_input("Press Enter to continue...")
            sheet["N%i"  % row].value = new_freq
            highlight_cell(sheet["N%i" % row])

       # Cell Measure
        if sheet["O%i"  % row].value is None:
            current_cellmeas = ""
        else:
            current_cellmeas = sheet["O%i"  % row].value
        new_cellmeas = sdq.get_cell_measures(uid)

        if str(current_cellmeas) != str(new_cellmeas):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_cellmeas=",current_cellmeas,"to new_cellmeas=",new_cellmeas
            #raw_input("Press Enter to continue...")
            sheet["O%i"  % row].value = new_cellmeas
            highlight_cell(sheet["O%i" % row])

       # Prov
        if sheet["P%i"  % row].value is None:
            current_prov = ""
        else:
            current_prov = sheet["P%i"  % row].value
        new_prov = sdq.get_prov(uid)

        if str(current_prov) != str(new_prov):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_prov=",current_prov,"to new_prov=",new_prov
            #raw_input("Press Enter to continue...")
            sheet["P%i"  % row].value = new_prov
            highlight_cell(sheet["P%i" % row])

       # ProvNote
        if sheet["Q%i"  % row].value is None:
            current_provn = ""
        else:
            current_provn = sheet["Q%i"  % row].value
        new_provn = sdq.get_provNote(uid)

        if str(current_provn) != str(new_provn):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_provn=",current_provn,"to new_provn=",new_provn
            #raw_input("Press Enter to continue...")
            sheet["Q%i"  % row].value = new_provn
            highlight_cell(sheet["Q%i" % row])

       # UID
        if sheet["S%i"  % row].value is None:
            current_uid = ""
        else:
            current_uid = sheet["S%i"  % row].value
        new_uid = sdq.get_uid(uid)

        if str(current_uid) != str(new_uid):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_uid=",current_uid,"to new_uid=",new_uid
            #raw_input("Press Enter to continue...")
            sheet["S%i"  % row].value = new_uid
            highlight_cell(sheet["S%i" % row])

       # VID
        if sheet["T%i"  % row].value is None:
            current_vid = ""
        else:
            current_vid = sheet["T%i"  % row].value
        new_vid = sdq.get_vid(uid)

        if str(current_vid) != str(new_vid):
            if first == 1 : print "uid=",uid ; first=0
            print "Correcting current_vid=",current_vid,"to new_vid=",new_vid
            #raw_input("Press Enter to continue...")
            sheet["T%i"  % row].value = new_vid
            highlight_cell(sheet["T%i" % row])

       # STID
       # current_stid = sheet["U%i"  % row].value
       # new_stid = sdq.get_stid(uid)

       # if str(current_stid != new_stid):
       #     sheet["U%i"  % row].value = new_stid
       #     highlight_cell(sheet["U%i" % row])

       # Mips Requesting
        if sheet["W%i"  % row].value is None:
            current_mipsrq = ""
        else:
            current_mipsrq = sheet["W%i"  % row].value
        new_mipsrq = sdq.get_mips_requesting(uid)

        if str(current_mipsrq) != str(new_mipsrq):
            #if first == 1 : print "uid=",uid ; first=0
            #print "Correcting current_mipsrq=",current_mipsrq,"to new_mipsrq=",new_mipsrq
            #raw_input("Press Enter to continue...")
            sheet["W%i"  % row].value = new_mipsrq
            highlight_cell(sheet["W%i" % row])

       # Template
       # current_XX = sheet["XX%i"  % row].value
       # new_XX = sdq.get_XX(uid)

       # if current_XX != new_XX:
       #     sheet["XX%i"  % row].value = new_XX
       #     highlight_cell(sheet["XX%i" % row])


workbook.save(OUTPUT_XLS)
